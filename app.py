from flask import (
    Flask, render_template, request,
    send_file, after_this_request,
    redirect, url_for, session, abort, jsonify
)
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import openpyxl
from openpyxl import load_workbook  
import datetime as dt
import tempfile
import os
import json
import time
from functools import wraps
from datetime import datetime

# ================= CONFIG =================
APP_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(APP_DIR, "מבנה.xlsx")
CONFIG_FILE = os.path.join(APP_DIR, "config.json")
STATE_FILE = os.path.join(APP_DIR, "state.json")
LOG_FILE = os.path.join(APP_DIR, "audit.log")
BOOT_FILE = os.path.join(APP_DIR, "boot.json")
TOUCH_LOG_PATH = os.path.join(APP_DIR, "data", "touch_log.json")

def load_config():
    with open(CONFIG_FILE, encoding="utf-8") as f:
        return json.load(f)

CONFIG = load_config()

app = Flask(__name__)
app.secret_key = CONFIG["secret_key"]
app.permanent_session_lifetime = dt.timedelta(hours=24)
EXPORT_YELLOW = "FFF2CC"
# ================= BOOT ID (invalidate sessions after server restart) =================
def _write_boot_id():
    boot = {"boot_id": f"{int(time.time())}"}
    try:
        with open(BOOT_FILE, "w", encoding="utf-8") as f:
            json.dump(boot, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    return boot["boot_id"]

def _read_boot_id():
    try:
        if os.path.exists(BOOT_FILE):
            with open(BOOT_FILE, encoding="utf-8") as f:
                data = json.load(f)
                if data.get("boot_id"):
                    return data["boot_id"]
    except Exception:
        pass
    return _write_boot_id()

CURRENT_BOOT_ID = _write_boot_id()

# ================= SESSION CHECK =================
@app.before_request
def check_login_timeout():
    if request.endpoint in ("login", "static"):
        return

    if "user" not in session:
        return redirect(url_for("login"))

    # force re-login after server restart
    if session.get("boot_id") != CURRENT_BOOT_ID:
        session.clear()
        return redirect(url_for("login"))

    login_at = session.get("login_at")
    if login_at:
        try:
            login_time = dt.datetime.fromisoformat(login_at)
            if dt.datetime.now() - login_time > dt.timedelta(hours=24):
                session.clear()
                return redirect(url_for("login"))
        except Exception:
            session.clear()
            return redirect(url_for("login"))

# ================= STYLES =================
YELLOW = PatternFill("solid", fgColor="FFF2CC")
TEAM_FILL = PatternFill("solid", fgColor="FFE699")
HEADER_FILL = PatternFill("solid", fgColor="E7E6E6")

THIN = Side(style="thin")
THICK = Side(style="thick")

BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_THICK = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
FONT_BOLD = Font(bold=True)

# ================= CONSTANTS =================
NAME_COL = 1
ID_COL = 2
HEADER_DATE_ROW = 4
HEADER_SHIFT_ROW = 5

# ✅ FIX: employees must start AFTER the shift header row (otherwise first employee “falls” into headers row)
EMP_START_ROW = 6

SHIFT_TYPES = ["בוקר", "ערב", "לילה"]

ACTIONS = [
    "השלמת שעות",
    "קיצור משמרת",
    "ביטול משמרת",
    "מחלה",
    "מילואים",
    "חופש",
]

# ================= HELPERS =================
def log_action(action, details=None):
    ts = dt.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    user = session.get("user", "anonymous")

    line = f"[{ts}] {user} | {action}"
    if details:
        line += " | " + " | ".join(details)

    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")
def find_first_employee_row(ws):
    """
    מחפש את השורה הראשונה שבה:
    - יש שם
    - ויש תעודת זהות
    - והיא מתחת לשורת המשמרות
    """
    row = HEADER_SHIFT_ROW + 1
    max_row = ws.max_row

    while row <= max_row:
        name = ws.cell(row, NAME_COL).value
        emp_id = ws.cell(row, ID_COL).value

        if name and emp_id:
            return row

        row += 1

    return HEADER_SHIFT_ROW + 1

def update_state(action_name=None):
    data = {
        "last_modified_by": session.get("user"),
        "last_modified_at": dt.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    }
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    if action_name:
        log_action(action_name)

def load_state():
    if not os.path.exists(STATE_FILE):
        return {"last_modified_by": "", "last_modified_at": ""}
    try:
        with open(STATE_FILE, encoding="utf-8") as f:
            data = json.load(f)
        return {
            "last_modified_by": data.get("last_modified_by", ""),
            "last_modified_at": data.get("last_modified_at", "")
        }
    except Exception:
        return {"last_modified_by": "", "last_modified_at": ""}

def apply_border(ws, sr, er, sc, ec, thick=False):
    border = BORDER_THICK if thick else BORDER_THIN
    for r in range(sr, er + 1):
        for c in range(sc, ec + 1):
            ws.cell(r, c).border = border

def daterange(start, end):
    d = start
    while d <= end:
        yield d
        d += dt.timedelta(days=1)

def safe_remove(path):
    try:
        os.remove(path)
    except Exception:
        pass

def read_audit_logs(max_lines=500):
    if not os.path.exists(LOG_FILE):
        return []
    try:
        with open(LOG_FILE, encoding="utf-8") as f:
            lines = f.readlines()[-max_lines:]
        logs = []
        for line in reversed(lines):
            line = line.strip()
            if not line:
                continue
            if line.startswith("[") and "]" in line and " | " in line:
                ts = line[1:line.index("]")]
                rest = line[line.index("]") + 1:].strip()
                if " | " in rest:
                    user, action = rest.split(" | ", 1)
                    logs.append({"ts": ts, "user": user.strip(), "action": action.strip()})
        return logs
    except Exception:
        return []

def save_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def _parse_date(value):
    return dt.datetime.strptime(value, "%Y-%m-%d").date()

def _clear_dynamic_columns_only(ws):
    """
    חשוב: לא עושים UNMERGE גורף!
    מפרקים רק מיזוגים שנוגעים בעמודות הדינמיות (C והלאה),
    ואז מוחקים רק את העמודות C..END כדי לבנות מחדש את טווח התאריכים.
    """
    try:
        merged = list(ws.merged_cells.ranges)
        for rng in merged:
            # אם המיזוג נוגע בעמודה 3 ומעלה (C..), מפרקים אותו
            if getattr(rng, "max_col", 0) >= 3:
                try:
                    ws.unmerge_cells(str(rng))
                except Exception:
                    pass
    except Exception:
        pass

    # מחיקת העמודות הדינמיות (C והלאה) בלבד
    if ws.max_column > 2:
        ws.delete_cols(3, ws.max_column - 2)

# ================= AUTH =================
def is_super_admin():
    return session.get("user") == CONFIG.get("super_admin")
def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper

def admin_required():
    if session.get("role") != "admin":
        abort(403)

# ================= LOAD TEAMS =================
def load_teams_with_rows():
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb.active

    teams = {}
    team_rows = {}
    current_team = None
    default_team = "מנהלי משמרת"

    row = find_first_employee_row(ws)
    empty = 0

    while empty < 10:

        name = ws.cell(row, NAME_COL).value
        emp_id = ws.cell(row, ID_COL).value
      
        if not name:
            empty += 1
            row += 1
            continue

        empty = 0
        name = str(name).strip()

# בדיקה אם השורה ממוזגת לרוחב → צוות
        is_team_row = any(
            rng.min_row == row and rng.max_row == row
            for rng in ws.merged_cells.ranges
        )

        if is_team_row:
            current_team = name
            teams.setdefault(current_team, [])
            team_rows[current_team] = row
        else:
            team_key = current_team or default_team
            teams.setdefault(team_key, [])
            teams[team_key].append({
                "name": name,
                "row": row
            })

        row += 1

    return teams, team_rows

def load_teams():
    teams, _ = load_teams_with_rows()
    return teams

# ================= HEADERS =================
def build_report_headers(ws, start_date, end_date):
    col = 3
    for d in daterange(start_date, end_date):
        start_col = col
        end_col = col + 2

        ws.merge_cells(
            start_row=HEADER_DATE_ROW,
            start_column=start_col,
            end_row=HEADER_DATE_ROW,
            end_column=end_col
        )

        dc = ws.cell(HEADER_DATE_ROW, start_col)
        dc.value = d
        dc.number_format = "DD.MM.YY"
        dc.font = FONT_BOLD
        dc.alignment = ALIGN_CENTER
        dc.fill = HEADER_FILL

        for i, shift in enumerate(SHIFT_TYPES):
            c = ws.cell(HEADER_SHIFT_ROW, start_col + i)
            c.value = f"משמרת {shift}"
            c.font = FONT_BOLD
            c.alignment = ALIGN_CENTER
            c.fill = HEADER_FILL
            c.border = BORDER_THIN

        apply_border(ws, HEADER_DATE_ROW, HEADER_SHIFT_ROW, start_col, end_col, thick=True)
        col += 3

# ================= ROUTES =================
@app.route("/login", methods=["GET", "POST"])
def login():
    global CONFIG
    if request.method == "POST":
        u = request.form.get("username")
        p = request.form.get("password")

        CONFIG = load_config()
        user = CONFIG["users"].get(u)

        if user and user.get("password") == p:
            session.clear()
            session["user"] = u
            session["role"] = user.get("role", "user")
            session["login_at"] = dt.datetime.now().isoformat()
            session["boot_id"] = CURRENT_BOOT_ID
            session.permanent = True
            log_action("login")
            return redirect(url_for("index"))

        return render_template("login.html", error="שם משתמש או סיסמה שגויים")

    return render_template("login.html")

@app.route("/logout")
def logout():
    log_action("logout")
    session.clear()
    return redirect(url_for("login"))

@app.route("/")
@login_required
def index():
    return render_template(
        "index.html",
        teams=load_teams(),
        actions=ACTIONS,
        shifts=[""] + SHIFT_TYPES,
        user=session["user"],
        role=session.get("role"),
        state=load_state(),
        config=CONFIG


    )

@app.get("/audit")
@login_required
def audit():
    admin_required()
    logs = read_audit_logs(max_lines=800)
    return render_template(
        "audit.html",
        logs=logs,
        user=session["user"],
        role=session.get("role"),
        state=load_state()
    )

@app.route("/users", methods=["GET"])
@login_required
def users():
    admin_required()
    cfg = load_config()

    return render_template(
        "users.html",
        users=cfg.get("users", {}),
        user=session["user"],
        role=session.get("role"),
        state=load_state()
    )

@app.post("/users/create")
@login_required
def users_create():
    admin_required()
    data = request.form or {}

    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    role = data.get("role", "user")

    cfg = load_config()

    if not username or not password:
        return render_template(
            "users.html",
            users=cfg.get("users", {}),
            error="חובה להזין שם משתמש וסיסמה",
            user=session["user"],
            role=session.get("role"),
            state=load_state()
        )

    if username in cfg["users"]:
        return render_template(
            "users.html",
            users=cfg.get("users", {}),
            error="שם משתמש כבר קיים",
            user=session["user"],
            role=session.get("role"),
            state=load_state()
        )

    if len(password) < 6:
        return render_template(
            "users.html",
            users=cfg.get("users", {}),
            error="סיסמה חייבת להכיל לפחות 6 תווים",
            user=session["user"],
            role=session.get("role"),
            state=load_state()
        )

    cfg["users"][username] = {
        "password": password,
        "role": role
    }

    save_config(cfg)
    update_state("create user")
    log_action("create user", [f"user={username}", f"role={role}"])

    # ✅ חזרה למסך ניהול משתמשים + רענון
    return redirect(url_for("users"))

@app.post("/users/update")
@login_required
def users_update():
    admin_required()
    data = request.json or {}

    username = data.get("username")
    password = data.get("password")
    role = data.get("role")

    cfg = load_config()
    user = cfg["users"].get(username)

    if not user:
        return jsonify(error="משתמש לא קיים"), 404

    if password:
        if len(password) < 6:
            return jsonify(error="סיסמה קצרה מדי"), 400
        user["password"] = password

    if role:
        user["role"] = role

    save_config(cfg)
    update_state("update user")
    log_action("update user", [f"user={username}"])

    return jsonify(ok=True)
@app.post("/users/delete")
@login_required
def users_delete():
    admin_required()
    data = request.json or {}

    username = data.get("username")
    password = data.get("password")

    current_user = session["user"]

    # ❌ מניעת מחיקה עצמית
    if username == current_user:
        return jsonify(error="לא ניתן למחוק את המשתמש שמחובר כעת"), 400

    cfg = load_config()

    # 🔐 בדיקת סיסמה של המשתמש המחובר
    admin_user = cfg["users"].get(current_user)
    if not admin_user or admin_user["password"] != password:
        return jsonify(error="סיסמה שגויה"), 403

    # בדיקה שהמשתמש הנמחק קיים
    if username not in cfg["users"]:
        return jsonify(error="משתמש לא קיים"), 404

    # ❗ לפחות מנהל אחד
    admins = [u for u in cfg["users"].values() if u["role"] == "admin"]
    if cfg["users"][username]["role"] == "admin" and len(admins) == 1:
        return jsonify(error="חייב להישאר לפחות מנהל אחד"), 400

    # מחיקה
    del cfg["users"][username]
    save_config(cfg)

    update_state("delete user")
    log_action(
        "delete user",
        [
            f"deleted={username}",
            f"by={current_user}"
        ]
    )

    return jsonify(ok=True)
@app.post("/reset")
@login_required
def reset():
    # 🔐 רק מנהל המערכת הראשי
    if session.get("user") != CONFIG.get("super_admin"):
        abort(403)

    if os.path.exists(STATE_FILE):
        os.remove(STATE_FILE)

    if os.path.exists(LOG_FILE):
        os.remove(LOG_FILE)

    if os.path.exists(PAYROLL_STATUS_PATH):
        os.remove(PAYROLL_STATUS_PATH)

    update_state("reset system")
    return jsonify({"ok": True})

@app.post("/touch")
@login_required
def touch():
    data = request.json or {}
    entries = data.get("entries", [])

    user = session.get("user")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # עדכון state
    state = {
        "last_modified_by": user,
        "last_modified_at": now
    }
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

    touch_log = load_touch_log()

    for e in entries:
        date = e.get("date")
        name = e.get("name")
        shift = e.get("shift") or ""
        key = f"{date}|{name}|{shift}"

        # 🔥 דריסה מלאה – אין היסטוריה
        touch_log[key] = {
            "touched_at": now,
            "by": user
        }

        log_action(
            "update entry",
            [
                f"employee={name}",
                f"date={date}",
                f"shift={e.get('shift')}",
                f"value={e.get('action')} {e.get('note') or ''}"
            ]
        )

    save_touch_log(touch_log)
    return jsonify(state)


@app.post("/export")
@login_required
def export():
    data = request.json or {}

    # --- validate input ---
    try:
        report_from = _parse_date(data["report_from"])
        report_to = _parse_date(data["report_to"])
    except Exception:
        return jsonify({"error": "report_from/report_to invalid. expected YYYY-MM-DD"}), 400

    if report_to < report_from:
        return jsonify({"error": "טווח תאריכים לא תקין: 'עד' קטן מ-'מדוח'"}), 400

    entries = data.get("entries", [])
    if not entries:
        return jsonify({"error": "no entries to export"}), 400

    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb.active

    # ✅ FIX: clean only dynamic columns (C..), without destroying static header merges
    _clear_dynamic_columns_only(ws)

    # build dynamic headers for the selected range
    build_report_headers(ws, report_from, report_to)

    teams, team_rows = load_teams_with_rows()
    emp_row = {e["name"]: e["row"] for t in teams.values() for e in t}

    # map (date, shift) -> column
    col_map = {}

    col = 3
    for d in daterange(report_from, report_to):
        for i, shift in enumerate(SHIFT_TYPES):
            col_map[(d, shift)] = col + i
        col += 3

    for e in entries:
        try:
            date = _parse_date(e["date"])
        except Exception:
            continue

        if not (report_from <= date <= report_to):
            continue

        name = e.get("name")
        shift = e.get("shift")
        action = (e.get("action") or "").strip()
        note = (e.get("note") or "").strip()

        if not name or not shift or name not in emp_row:
            continue

        col = col_map.get((date, shift))
        if not col:
            continue

        text = action
        if note:
            text = f"{action} – {note}" if action else note

        


        cell = ws.cell(emp_row[name], col)
        cell.value = text
        cell.fill = YELLOW
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_CENTER

    # global borders
    apply_border(ws, 1, ws.max_row, 1, ws.max_column)

    # team header rows (merge across current max_column)
    for team, r in team_rows.items():
        try:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ws.max_column)
        except Exception:
            pass
        c = ws.cell(r, 1)
        c.value = team
        c.fill = TEAM_FILL
        c.font = FONT_BOLD
        c.alignment = ALIGN_CENTER
        apply_border(ws, r, r, 1, ws.max_column, thick=True)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    path = tmp.name
    tmp.close()
    wb.save(path)

    # ✅ no duplicate "export excel" lines (update_state already logs action_name)
    update_state("export excel")

    @after_this_request
    def cleanup(resp):
        safe_remove(path)
        return resp

    return send_file(path, as_attachment=True, download_name="hours_report.xlsx")


def is_marked_as_done(cell) -> bool:
    fill = getattr(cell, "fill", None)
    if not fill or fill.patternType != "solid":
        return False

    sc = fill.start_color
    if not sc or not sc.rgb:
        return False

    # ⭐️ זה התיקון הקריטי
    rgb = str(sc.rgb).upper()

    # אם הצבע שונה מהצהוב שהמערכת ייצרה → נחשב "טופל"
    return rgb != EXPORT_YELLOW



# 🎯 צבעים שמותרים כ"כתום שכר" בלבד
PAYROLL_ORANGE_RGB = {
    "FFFFC000",  # Excel Orange
    "FFFFA500",  # Orange
    "FFED7D31",  # Custom orange
}

def is_payroll_done_cell(cell) -> bool:
    fill = cell.fill
    if not fill or fill.patternType != "solid":
        return False

    # חייב להיות ערך בתא
    if not cell.value or not str(cell.value).strip():
        return False

    fg = fill.fgColor
    if not fg:
        return False

    # 🟨 אם זה בדיוק הצהוב של המערכת → זה לא שכר
    if fg.type == "rgb" and fg.rgb:
        rgb = fg.rgb.upper()
        if len(rgb) == 8 and rgb.startswith("00"):
            rgb = "FF" + rgb[2:]

        if rgb == "FFFFF2CC":  # הצהוב שלך
            return False

    # ✅ כל צבע אחר (כולל THEME) נחשב שכר
    return True

def _to_date_iso(v):
    # header date cell יכול להיות date/datetime/str
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, str):
        v = v.strip()
        # אם אצלך מופיע "01.01.25" או דומה - אפשר להרחיב פה
        try:
            # אם זה מגיע כ-YYYY-MM-DD
            return dt.datetime.strptime(v, "%Y-%m-%d").date().isoformat()
        except Exception:
            pass
    return None

def build_col_meta_from_export(ws):
    """
    קורא את האקסל 'המטריצי' שיוצא מהמערכת:
    - תאריך בשורה HEADER_DATE_ROW (4) עם מיזוגים
    - משמרת בשורה HEADER_SHIFT_ROW (5)
    מחזיר dict: {col_index: (date_iso, shift_name)}
    """
    meta = {}

    for col in range(3, ws.max_column + 1):
        # date: לפעמים הערך נמצא רק בעמודה הראשונה של המיזוג
        dv = ws.cell(HEADER_DATE_ROW, col).value
        if not dv:
            # נלך שמאלה עד שנמצא ערך (כדי להתמודד עם merged header)
            c = col
            while c >= 3 and not ws.cell(HEADER_DATE_ROW, c).value:
                c -= 1
            dv = ws.cell(HEADER_DATE_ROW, c).value if c >= 3 else None

        date_iso = _to_date_iso(dv)
        if not date_iso:
            continue

        sv = ws.cell(HEADER_SHIFT_ROW, col).value or ""
        sv = str(sv).strip()
        # מצפה ל"משמרת בוקר" / "משמרת ערב" / "משמרת לילה"
        shift = sv.replace("משמרת", "").replace("\xa0", " ").strip() if sv else ""
        if not shift:
            continue

        meta[col] = (date_iso, shift)

    return meta

def is_team_row(ws, row):
    # אם השורה ממוזגת לרוחב => זו שורת צוות
    try:
        return any(rng.min_row == row and rng.max_row == row for rng in ws.merged_cells.ranges)
    except Exception:
        return False
PAYROLL_STATUS_PATH = os.path.join(APP_DIR, "data", "payroll_status.json")

def load_payroll_status():
    if not os.path.exists(PAYROLL_STATUS_PATH):
        return {}
    with open(PAYROLL_STATUS_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def save_payroll_status(data):
    os.makedirs(os.path.dirname(PAYROLL_STATUS_PATH), exist_ok=True)
    with open(PAYROLL_STATUS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
def load_touch_log():
    if not os.path.exists(TOUCH_LOG_PATH):
        return {}
    with open(TOUCH_LOG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def save_touch_log(data):
    os.makedirs(os.path.dirname(TOUCH_LOG_PATH), exist_ok=True)
    with open(TOUCH_LOG_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def cell_fill_debug(cell) -> dict:
    """
    דיבאג בטוח ל־fill — בלי indexed / theme שגורמים לשגיאות
    """
    try:
        f = cell.fill
        if not f:
            return {"fill": None}

        fg = f.fgColor
        return {
            "patternType": f.patternType,
            "fgType": fg.type if fg else None,
            "fgRgb": fg.rgb if fg and fg.type == "rgb" else None,
        }
    except Exception as ex:
        return {"error": str(ex)}





@app.route("/upload-payroll", methods=["POST"])
@login_required
def upload_payroll():
    if session.get("role") != "admin":
        return jsonify({"error": "forbidden"}), 403

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "missing file"}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=False)
        ws = wb.active
    except Exception as ex:
        return jsonify({"error": f"failed to read xlsx: {ex}"}), 400

    # col -> (date_iso, shift)
    col_meta = build_col_meta_from_export(ws)

    payroll = load_payroll_status()
    now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    by = session.get("user", "admin")

    updated = 0
    scanned_cells = 0
    marked_cells = 0
    sample_printed = 0

    # עובדים החל מ-EMP_START_ROW
    for r in range(EMP_START_ROW, ws.max_row + 1):
        if is_team_row(ws, r):
            continue

        name = ws.cell(r, NAME_COL).value
        emp_id = ws.cell(r, ID_COL).value

        if not name or not emp_id:
            continue

        name = str(name).strip()

        # עובר רק על עמודות דינמיות שיש להן meta (תאריך+משמרת)
        for col, (date_iso, shift) in col_meta.items():
            cell = ws.cell(r, col)
            scanned_cells += 1

            dbg = cell_fill_debug(cell)

            print(
                f"[PAYROLL SCAN] r={r} c={col} name={name} "
                f"date={date_iso} shift={shift} fill={dbg}"
            )

            if is_payroll_done_cell(cell):
                key = f"{date_iso}|{name}|{shift}"

                # ⛔ אם כבר טופל – לא נוגעים
                if payroll.get(key, {}).get("done"):
                    continue

                payroll[key] = {
                    "done": True,
                    "updated_at": now,
                    "by": by
                }
                updated += 1

                # דיבאג: נדפיס כמה דוגמאות ראשונות כדי לוודא שאנחנו תופסים צבעים
                if sample_printed < 8:
                    dbg = cell_fill_debug(cell)
                    print(f"[PAYROLL DEBUG] r={r} c={col} name={name} date={date_iso} shift={shift} fill={dbg}")
                    sample_printed += 1

    save_payroll_status(payroll)
    log_action("upload payroll", [f"updated={updated}"])
    update_state("upload payroll")

    # דיבאג מסכם
    print("PAYROLL DEBUG SUMMARY:",
          "col_meta_cols=", len(col_meta),
          "scanned_cells=", scanned_cells,
          "marked_cells=", marked_cells,
          "updated_new=", updated,
          "TOTAL_PAYROLL_KEYS=", len(payroll))

    return jsonify({
        "ok": True,
        "updated": updated,
        "total_keys": len(payroll),
        "scanned_cells": scanned_cells,
        "marked_cells": marked_cells,
        "meta_cols": len(col_meta),
    })

@app.route("/payroll-status", methods=["GET"])
@login_required
def payroll_status():
    data = load_payroll_status()
    return jsonify(data)


@app.get("/payroll-dirty")
@login_required
def payroll_dirty():
    payroll = load_payroll_status()
    touch_log = load_touch_log()

    dirty = {}

    for key, p in payroll.items():
        t = touch_log.get(key)
        if not t:
            continue

        # 🔥 בדיקה נקייה: אותו key בלבד
        if t["touched_at"] > p["updated_at"]:
            dirty[key] = {
                "payroll_at": p["updated_at"],
                "touched_at": t["touched_at"],
                "by": t["by"]
            }

    return jsonify(dirty)

if __name__ == "__main__":
    app.run(debug=True)